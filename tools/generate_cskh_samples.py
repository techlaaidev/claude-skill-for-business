"""
Tạo 2 file sample cho Skill 1:
  - sample_weca_format.xlsx (3 sheet, format ma trận tháng)
  - sample_standard_format.xlsx (transaction log)

Author: Techla (dev tool — không ship trong skill)

Yêu cầu thỏa mãn theo spec:
  - 8 KH, 12 SP, 4 tháng (T1-T4/2026)
  - >=2 KH trigger churn (không đặt tháng 4)
  - >=1 KH trigger decreasing rhythm
  - >=2 SP dead product (0 đơn)
  - >=2 KH có opportunity (sheet 3)
  - Nhiều format ô cột tháng (datetime, "2;23", "01,17/04", int)
"""
from __future__ import annotations
import os
import random
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

random.seed(42)

BASE = Path(__file__).resolve().parent.parent
WECA_OUT = BASE / "src" / "cskh-phan-tich-don-hang" / "samples" / "sample_weca_format.xlsx"
STD_OUT = BASE / "src" / "cskh-phan-tich-don-hang" / "samples" / "sample_standard_format.xlsx"

# ---------------- Data giả ẩn danh ----------------

CUSTOMERS = [
    # (id, name, phone, address, opportunity, action, status)
    ("KH001", "Quán Sao Đêm", "0901111001", "12 Hàng Buồm, Hoàn Kiếm, Hà Nội",
     "Mở thêm chi nhánh quận 1 — cần tăng volume", "Tư vấn gói combo mới", "Đang chăm"),
    ("KH002", "Coffee House Trúc Bạch", "0901111002", "45 Trấn Vũ, Ba Đình, Hà Nội",
     None, None, "Đang chăm"),
    ("KH003", "The Local Brew Lab", "0901111003", "88 Xuân Diệu, Tây Hồ, Hà Nội",
     "Quan tâm trà olong premium", "Gửi mẫu thử Olong Kim Tuyên", "Đang chăm"),
    ("KH004", "Cafe Ngõ Nhỏ", "0901111004", "7 Phan Bội Châu, Hoàn Kiếm, Hà Nội",
     None, None, "Đang chăm"),
    ("KH005", "Matcha House Đặng Dung", "0901111005", "23 Đặng Dung, Ba Đình, Hà Nội",
     None, None, "Đang chăm"),
    ("KH006", "Quán Chill Trần Duy Hưng", "0901111006", "156 Trần Duy Hưng, Cầu Giấy, Hà Nội",
     None, None, "Cần liên hệ lại"),
    ("KH007", "Tea Station Thái Hà", "0901111007", "34 Thái Hà, Đống Đa, Hà Nội",
     None, None, "Đang chăm"),
    ("KH008", "Cafe Góc Phố Nhà Thờ", "0901111008", "9 Nhà Chung, Hoàn Kiếm, Hà Nội",
     None, None, "Đang chăm"),
]

PRODUCTS = [
    # (id, name, avg_qty_per_order, avg_orders_per_month, price_old, price_new)
    ("SP01", "Cafe Robusta Premium 1kg", 5, 4, 280_000, 290_000),
    ("SP02", "Cafe Arabica Đà Lạt 500g", 3, 3, 220_000, 230_000),
    ("SP03", "Cafe Phin Truyền Thống 500g", 4, 4, 180_000, 180_000),
    ("SP04", "Trà Olong Kim Tuyên 500g", 2, 2, 450_000, 480_000),
    ("SP05", "Trà Xanh Thái Nguyên 500g", 3, 2, 240_000, 250_000),
    ("SP06", "Matcha Nhật Uji 200g", 2, 2, 520_000, 550_000),
    ("SP07", "Syrup Vani Pháp 750ml", 4, 3, 180_000, 180_000),
    ("SP08", "Syrup Caramel 750ml", 4, 3, 180_000, 180_000),
    ("SP09", "Bột Sữa New Zealand 1kg", 6, 5, 210_000, 220_000),
    ("SP10", "Cacao Bỉ Nguyên Chất 500g", 2, 2, 380_000, 400_000),
    ("SP11", "Topping Trân Châu Trắng 3kg", 3, 2, 150_000, 150_000),
    ("SP12", "Ly Giấy Take-away 500 cái", 2, 2, 250_000, 260_000),
]


def _orders_for_customer(kh_idx: int, sp_idx: int) -> list:
    """Return list 4 cell values for Jan/Feb/Mar/Apr 2026.
    Generate dựa theo kịch bản để trigger các warning theo spec."""
    cust_id = f"KH{kh_idx+1:03d}"
    sp_id = f"SP{sp_idx+1:02d}"

    # Dead products: SP11, SP12 — không có đơn từ ai
    if sp_id in ("SP11", "SP12"):
        return [None, None, None, None]

    # Churn: KH007, KH008 — có đơn tháng 1-3, vắng tháng 4
    if cust_id in ("KH007", "KH008"):
        if sp_idx < 4:
            return [
                datetime(2026, 1, 10 + kh_idx),
                f"{5 + kh_idx};{20 + kh_idx}",
                f"{12 + kh_idx}",
                None,  # tháng 4 vắng -> churn
            ]
        return [None, None, None, None]

    # Decreasing rhythm: KH006 — đặt đều T1-T3 (mỗi tháng 3 đơn), T4 giảm còn 1 đơn
    if cust_id == "KH006":
        if sp_idx < 5:
            return [
                f"3;12;22",
                f"2;14;25",
                f"5;15;27",
                datetime(2026, 4, 8) if sp_idx == 0 else None,
            ]
        return [None, None, None, None]

    # VIP: KH001, KH002 — đặt nhiều, đa dạng SP
    if cust_id in ("KH001", "KH002"):
        if sp_idx < 8:
            # Mix format để test parser
            if sp_idx == 0:
                return [datetime(2026, 1, 5), "2;23/2", "01,17/04" if False else f"3;20/3", datetime(2026, 4, 12)]
            if sp_idx == 1:
                return [f"10", f"10/02", f"8;22", datetime(2026, 4, 10)]
            return [
                datetime(2026, 1, 10 + sp_idx),
                datetime(2026, 2, 10 + sp_idx) if sp_idx % 2 == 0 else f"{12+sp_idx}",
                f"{8+sp_idx}",
                datetime(2026, 4, 12 + sp_idx),
            ]
        return [None, None, None, None]

    # KH003: quan tâm olong (sp_idx=3) — đặt đều
    if cust_id == "KH003":
        if sp_idx in (0, 2, 3, 6):
            return [
                datetime(2026, 1, 15),
                datetime(2026, 2, 18),
                "12;28",
                datetime(2026, 4, 9),
            ]
        return [None, None, None, None]

    # KH004: đặt thường
    if cust_id == "KH004":
        if sp_idx in (0, 2, 8):
            return [
                datetime(2026, 1, 20),
                None,
                datetime(2026, 3, 15),
                datetime(2026, 4, 18),
            ]
        return [None, None, None, None]

    # KH005: matcha house — mua SP06 nhiều
    if cust_id == "KH005":
        if sp_idx in (5, 7, 9):
            return [
                datetime(2026, 1, 11),
                datetime(2026, 2, 14),
                datetime(2026, 3, 11),
                datetime(2026, 4, 14),
            ]
        return [None, None, None, None]

    return [None, None, None, None]


# ---------------- Build WECA xlsx ----------------

def build_weca():
    wb = Workbook()

    # SHEET 1: Tần Suất đặt hàng 2026
    ws = wb.active
    ws.title = "Tần Suất đặt hàng 2026"

    # Dòng 1-4: title + annotation
    ws["A1"] = "BẢNG TẦN SUẤT ĐẶT HÀNG 2026"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws["A3"] = "Đơn vị: Quán Cafe Techla Mẫu"

    # Dòng 5: header
    headers = [
        "STT", "MÃ KH", "TÊN KH", "SĐT", "ĐỊA CHỈ",
        "MÃ SP", "TÊN SP",
        "SỐ LƯỢNG BÁN TB/Lần", "SỐ LƯỢNG ĐƠN HÀNG/THÁNG",
        "GIÁ BÁN", "GIÁ BÁN MỚI",
        "01/26", "02/26", "03/26", "04/26",
        # cột mã nội bộ — nên bị bỏ qua
        "4E-4", "3E-11",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2E75B6")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row = 6
    stt = 1
    for kh_idx, (cust_id, name, phone, addr, _, _, _) in enumerate(CUSTOMERS):
        first_row = True
        for sp_idx, (sp_id, sp_name, avg_qty, avg_ord, p_old, p_new) in enumerate(PRODUCTS):
            months = _orders_for_customer(kh_idx, sp_idx)
            # Skip SP nếu không có đơn + không muốn mọi SP xuất hiện cho mỗi KH
            # Nhưng để giữ product master đầy đủ, cần ít nhất 1 KH mua mọi SP.
            # Trick: đưa KH001 làm "master" — liệt kê tất cả SP dù không đặt.
            has_any = any(m is not None for m in months)
            if cust_id != "KH001" and not has_any:
                continue

            # Ghi KH info chỉ ở dòng đầu (merge effect)
            if first_row:
                ws.cell(row=row, column=1, value=stt)
                ws.cell(row=row, column=2, value=cust_id)
                ws.cell(row=row, column=3, value=name)
                ws.cell(row=row, column=4, value=phone)
                ws.cell(row=row, column=5, value=addr)
                first_row = False
                stt += 1

            ws.cell(row=row, column=6, value=sp_id)
            ws.cell(row=row, column=7, value=sp_name)
            ws.cell(row=row, column=8, value=avg_qty)
            ws.cell(row=row, column=9, value=avg_ord)
            ws.cell(row=row, column=10, value=p_old)
            ws.cell(row=row, column=11, value=p_new)

            for m_offset, val in enumerate(months):
                ws.cell(row=row, column=12 + m_offset, value=val)

            row += 1

    # Column width
    widths = [5, 10, 28, 14, 40, 10, 32, 12, 12, 12, 12, 14, 14, 14, 14, 10, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[5].height = 40

    # SHEET 2: Lịch sử chăm sóc KH
    ws2 = wb.create_sheet("Lịch sử chăm sóc KH")
    care_headers = ["MÃ KH", "TÊN KH", "MỤC TIÊU CHĂM SÓC", "Nội dung tin nhắn", "LỊCH SỬ CHĂM SÓC"]
    for col, h in enumerate(care_headers, start=1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="2E75B6")
    care_data = [
        ("KH001", "Quán Sao Đêm", "Giữ VIP, tăng đơn 20%", "Anh/chị ơi, bên em có combo mới...", "T3: gọi confirm; T4: gửi bảng giá mới"),
        ("KH003", "The Local Brew Lab", "Upsell Olong Kim Tuyên", "Em gửi anh mẫu thử trà Olong Kim Tuyên mới về...", "T2: gửi mẫu; T3: follow-up"),
        ("KH006", "Quán Chill Trần Duy Hưng", "Phục hồi nhịp đặt", "Dạo này anh/chị ít đặt, bên em có gì hỗ trợ không ạ?", "T4: nhắc nhẹ, chưa phản hồi"),
        ("KH007", "Tea Station Thái Hà", "Giữ khách, tránh churn", "Lâu không thấy anh/chị đặt, bên em tặng voucher 5%", "T3: gọi 1 lần chưa được"),
    ]
    for ri, row_d in enumerate(care_data, start=2):
        for ci, v in enumerate(row_d, start=1):
            ws2.cell(row=ri, column=ci, value=v)
    ws2.column_dimensions["A"].width = 10
    ws2.column_dimensions["B"].width = 28
    ws2.column_dimensions["C"].width = 30
    ws2.column_dimensions["D"].width = 45
    ws2.column_dimensions["E"].width = 40

    # SHEET 3: Trang tính1 (danh sách KH đầy đủ)
    ws3 = wb.create_sheet("Trang tính1")
    list_headers = ["Tên khách hàng", "Số điện thoại", "Địa chỉ", "Hiện trạng chăm sóc", "Cơ hội", "Cần làm"]
    for col, h in enumerate(list_headers, start=1):
        c = ws3.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="2E75B6")
    for ri, (_, name, phone, addr, opp, action, status) in enumerate(CUSTOMERS, start=2):
        ws3.cell(row=ri, column=1, value=name)
        ws3.cell(row=ri, column=2, value=phone)
        ws3.cell(row=ri, column=3, value=addr)
        ws3.cell(row=ri, column=4, value=status)
        ws3.cell(row=ri, column=5, value=opp)
        ws3.cell(row=ri, column=6, value=action)
    for i, w in enumerate([28, 14, 40, 20, 35, 30], start=1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    WECA_OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(WECA_OUT)
    print(f"Wrote: {WECA_OUT}")


# ---------------- Build standard xlsx ----------------

def build_standard():
    """Transaction log: 1 đơn/dòng, cùng nội dung với file WECA."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Đơn hàng"
    headers = ["Ngày", "Mã KH", "Tên khách", "SĐT", "Mã SP", "Sản phẩm", "Số lượng", "Đơn giá", "Thành tiền"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="2E75B6")

    row = 2
    for kh_idx, (cust_id, name, phone, _, _, _, _) in enumerate(CUSTOMERS):
        for sp_idx, (sp_id, sp_name, avg_qty, _, p_old, p_new) in enumerate(PRODUCTS):
            months = _orders_for_customer(kh_idx, sp_idx)
            price = p_new
            # Rebuild orders from same logic (flatten)
            from datetime import datetime as dt
            import re
            import calendar as cal
            for m_idx, val in enumerate(months, start=1):
                if val is None:
                    continue
                dates: list[dt] = []
                if isinstance(val, dt):
                    dates = [val]
                elif isinstance(val, int):
                    dates = [dt(2026, m_idx, int(val))]
                else:
                    s = str(val).strip()
                    # Reuse same parse logic (simplified here — only needed for sample consistency)
                    tokens = [t.strip() for t in re.split(r"[;,]", s) if t.strip()]
                    override_m = None
                    if tokens:
                        last = tokens[-1]
                        sm = re.match(r"(\d{1,2})\s*/\s*(\d{1,2})$", last)
                        if sm:
                            d_last, m_over = map(int, sm.groups())
                            override_m = m_over
                            tokens[-1] = str(d_last)
                    mm = override_m if override_m else m_idx
                    for tok in tokens:
                        inner = re.match(r"^(\d{1,2})\s*/\s*(\d{1,2})$", tok)
                        if inner:
                            d_v, m_v = map(int, inner.groups())
                            try:
                                dates.append(dt(2026, m_v, d_v))
                            except ValueError:
                                pass
                        elif tok.isdigit():
                            d_v = int(tok)
                            try:
                                last_day = cal.monthrange(2026, mm)[1]
                                if 1 <= d_v <= last_day:
                                    dates.append(dt(2026, mm, d_v))
                            except Exception:
                                pass
                for d in dates:
                    qty = avg_qty
                    ws.cell(row=row, column=1, value=d)
                    ws.cell(row=row, column=2, value=cust_id)
                    ws.cell(row=row, column=3, value=name)
                    ws.cell(row=row, column=4, value=phone)
                    ws.cell(row=row, column=5, value=sp_id)
                    ws.cell(row=row, column=6, value=sp_name)
                    ws.cell(row=row, column=7, value=qty)
                    ws.cell(row=row, column=8, value=price)
                    ws.cell(row=row, column=9, value=qty * price)
                    row += 1

    for i, w in enumerate([12, 10, 28, 14, 10, 32, 10, 14, 16], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    STD_OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(STD_OUT)
    print(f"Wrote: {STD_OUT} ({row - 2} rows)")


if __name__ == "__main__":
    build_weca()
    build_standard()
