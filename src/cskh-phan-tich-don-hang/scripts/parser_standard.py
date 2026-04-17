"""
Parser format chuẩn (transaction log): 1 dòng = 1 đơn.
Author: Techla — v1.0.0 — License: xem LICENSE.md

Usage:
    python scripts/parser_standard.py input.xlsx --output parsed.json
"""
from __future__ import annotations
import os
import sys
import argparse
from pathlib import Path
from datetime import datetime, date

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import load_workbook
from parser_utils import normalize_header, to_float, safe_str, build_column_map
from schema import Customer, Product, OrderEvent, ParsedData, now_iso


def _cell_to_iso_date(v) -> str | None:
    """Cast cell value thành ISO date string."""
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    s = str(v).strip()
    # dd/mm/yyyy or dd-mm-yyyy
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y", "%d.%m.%Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def parse_standard_file(path: str) -> ParsedData:
    wb = load_workbook(path, data_only=True)
    ws = wb.active  # Default first sheet
    warnings: list[str] = []

    if ws.max_row < 2:
        raise ValueError("File rỗng hoặc chỉ có header.")

    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    col_map = build_column_map(headers)

    required = ["date", "product_name", "qty"]
    missing = [f for f in required if f not in col_map]
    if missing:
        raise ValueError(
            f"Thiếu cột bắt buộc: {missing}. "
            "Cần có ít nhất: ngày đặt + tên SP + số lượng. "
            "File của bạn có các cột: " + ", ".join(str(h) for h in headers if h)
        )

    customers: dict[str, Customer] = {}
    products: dict[str, Product] = {}
    orders: list[OrderEvent] = []

    years_seen: set[int] = set()

    # Tính unit_price trung bình cho mỗi SP (cuối function sẽ set)
    product_price_samples: dict[str, list[float]] = {}
    product_qty_samples: dict[str, list[float]] = {}

    for row_idx in range(2, ws.max_row + 1):
        row = [ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)]
        if all(c is None for c in row):
            continue

        iso_date = _cell_to_iso_date(row[col_map["date"]])
        if iso_date is None:
            warnings.append(f"Dòng {row_idx}: ngày không hợp lệ, bỏ qua.")
            continue
        years_seen.add(int(iso_date[:4]))

        # Customer
        cust_name = safe_str(row[col_map["customer_name"]]) if "customer_name" in col_map else None
        cust_id = safe_str(row[col_map["customer_id"]]) if "customer_id" in col_map else None
        phone = safe_str(row[col_map["phone"]]) if "phone" in col_map else None
        address = safe_str(row[col_map["address"]]) if "address" in col_map else None

        customer_key = cust_id or cust_name or phone
        if not customer_key:
            warnings.append(f"Dòng {row_idx}: không xác định được khách hàng, bỏ qua.")
            continue

        if customer_key not in customers:
            customers[customer_key] = Customer(
                customer_id=customer_key,
                name=cust_name or cust_id or customer_key,
                phone=phone,
                address=address,
            )
        else:
            c = customers[customer_key]
            if phone and not c.phone: c.phone = phone
            if address and not c.address: c.address = address

        # Product
        prod_name = safe_str(row[col_map["product_name"]])
        prod_id = safe_str(row[col_map["product_id"]]) if "product_id" in col_map else None
        product_key = prod_id or prod_name
        if not product_key:
            warnings.append(f"Dòng {row_idx}: thiếu tên/mã SP, bỏ qua.")
            continue

        qty = to_float(row[col_map["qty"]])
        unit_price = to_float(row[col_map["unit_price"]]) if "unit_price" in col_map else None
        revenue = to_float(row[col_map["revenue"]]) if "revenue" in col_map else None
        if revenue is None and qty and unit_price:
            revenue = qty * unit_price
        if unit_price is None and qty and revenue and qty != 0:
            unit_price = revenue / qty

        if product_key not in products:
            products[product_key] = Product(
                product_id=product_key,
                name=prod_name or prod_id or product_key,
                unit_price=unit_price,
            )
        product_price_samples.setdefault(product_key, [])
        product_qty_samples.setdefault(product_key, [])
        if unit_price:
            product_price_samples[product_key].append(unit_price)
        if qty:
            product_qty_samples[product_key].append(qty)

        orders.append(OrderEvent(
            customer_id=customer_key,
            product_id=product_key,
            order_date=iso_date,
            qty=qty,
            estimated_revenue=revenue,
        ))

    # Tính avg price/qty cho mỗi SP
    for pid, p in products.items():
        prices = product_price_samples.get(pid, [])
        qtys = product_qty_samples.get(pid, [])
        if prices and p.unit_price is None:
            p.unit_price = sum(prices) / len(prices)
        if qtys:
            p.avg_qty_per_order = sum(qtys) / len(qtys)

    year = max(years_seen) if years_seen else datetime.now().year

    return ParsedData(
        source_file=os.path.basename(path),
        parsed_at=now_iso(),
        format="standard",
        year=year,
        customers=list(customers.values()),
        products=list(products.values()),
        orders=orders,
        warnings=warnings,
    )


def main():
    ap = argparse.ArgumentParser(description="Parser transaction log chuẩn -> JSON.")
    ap.add_argument("input", help="Đường dẫn file .xlsx")
    ap.add_argument("--output", "-o", default="parsed.json")
    args = ap.parse_args()

    if not Path(args.input).exists():
        print(f"Lỗi: không tìm thấy file '{args.input}'.", file=sys.stderr)
        sys.exit(1)

    try:
        parsed = parse_standard_file(args.input)
    except Exception as e:
        print(f"Lỗi khi parse: {e}", file=sys.stderr)
        sys.exit(2)

    parsed.save_json(args.output)
    print(f"OK: {len(parsed.customers)} KH, {len(parsed.products)} SP, {len(parsed.orders)} đơn -> {args.output}")
    if parsed.warnings:
        print("Cảnh báo:")
        for w in parsed.warnings[:10]:
            print(f"  - {w}")


if __name__ == "__main__":
    main()
