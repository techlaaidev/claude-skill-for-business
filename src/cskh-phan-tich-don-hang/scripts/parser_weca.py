"""
Parser WECA: đọc file Excel bảng ma trận tháng -> JSON chuẩn.
Author: Techla — v1.0.0 — License: xem LICENSE.md

Usage:
    python scripts/parser_weca.py input.xlsx --output parsed.json
"""
from __future__ import annotations
import os
import sys
import argparse
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import load_workbook
from parser_utils import normalize_header, to_float, safe_str
from weca_sheet_classifier import classify_sheet, find_header_row_main
from weca_date_cell_parser import parse_month_cell, parse_column_header_to_month
from schema import Customer, Product, OrderEvent, ParsedData, now_iso


# Các cột cố định của sheet main WECA
MAIN_FIXED_COLS = {
    "stt": "stt",
    "ma kh": "customer_id",
    "ten kh": "name",
    "sdt": "phone",
    "dia chi": "address",
    "ma sp": "product_id",
    "ten sp": "product_name",
    "so luong ban tb/lan": "avg_qty_per_order",
    "so luong don hang/thang": "avg_orders_per_month",
    "gia ban": "unit_price",
    "gia ban moi": "unit_price_new",
}


def _map_main_headers(header_row: list) -> dict[str, int]:
    """Map header -> col index cho sheet main."""
    out: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        key = normalize_header(cell)
        if key in MAIN_FIXED_COLS:
            out[MAIN_FIXED_COLS[key]] = idx
    return out


def _find_month_cols(header_row: list) -> list[tuple[int, int, int]]:
    """Trả list (col_idx, month, year) cho các cột tháng MM/YY."""
    cols = []
    for idx, cell in enumerate(header_row):
        parsed = parse_column_header_to_month(cell)
        if parsed:
            cols.append((idx, parsed[0], parsed[1]))
    return cols


def parse_main_sheet(ws, warnings: list[str]):
    """Parse sheet chính. Trả (customers_dict, products_dict, orders_list, year)."""
    header_row_idx = find_header_row_main(ws) or 5
    header_row = [ws.cell(row=header_row_idx, column=c).value for c in range(1, ws.max_column + 1)]

    fixed = _map_main_headers(header_row)
    month_cols = _find_month_cols(header_row)

    if not month_cols:
        warnings.append("Không tìm thấy cột tháng MM/YY trong sheet chính.")
        year = 2026
    else:
        year = month_cols[0][2]

    if "customer_id" not in fixed or "product_id" not in fixed:
        warnings.append("Sheet chính thiếu cột MÃ KH hoặc MÃ SP.")

    customers: dict[str, Customer] = {}
    products: dict[str, Product] = {}
    orders: list[OrderEvent] = []

    # Forward-fill state
    last_customer_id = None
    last_name = None
    last_phone = None
    last_address = None

    for row_idx in range(header_row_idx + 1, ws.max_row + 1):
        row = [ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)]
        # Skip empty row
        if all(c is None for c in row):
            continue

        # Extract customer (forward-fill)
        ma_kh = safe_str(row[fixed["customer_id"]]) if "customer_id" in fixed else None
        ten_kh = safe_str(row[fixed["name"]]) if "name" in fixed else None
        sdt = safe_str(row[fixed["phone"]]) if "phone" in fixed else None
        addr = safe_str(row[fixed["address"]]) if "address" in fixed else None

        if ma_kh:
            last_customer_id = ma_kh
        if ten_kh:
            last_name = ten_kh
        if sdt:
            last_phone = sdt
        if addr:
            last_address = addr

        customer_id = last_customer_id
        if not customer_id:
            continue

        if customer_id not in customers:
            customers[customer_id] = Customer(
                customer_id=customer_id,
                name=last_name or customer_id,
                phone=last_phone,
                address=last_address,
            )
        else:
            # Update nếu giá trị mới rõ ràng hơn
            c = customers[customer_id]
            if last_name and (not c.name or c.name == c.customer_id):
                c.name = last_name
            if last_phone and not c.phone:
                c.phone = last_phone
            if last_address and not c.address:
                c.address = last_address

        # Extract product
        product_id = safe_str(row[fixed["product_id"]]) if "product_id" in fixed else None
        product_name = safe_str(row[fixed["product_name"]]) if "product_name" in fixed else None
        if not product_id:
            continue

        price_new = to_float(row[fixed["unit_price_new"]]) if "unit_price_new" in fixed else None
        price_old = to_float(row[fixed["unit_price"]]) if "unit_price" in fixed else None
        unit_price = price_new if price_new else price_old
        avg_qty = to_float(row[fixed["avg_qty_per_order"]]) if "avg_qty_per_order" in fixed else None
        avg_orders = to_float(row[fixed["avg_orders_per_month"]]) if "avg_orders_per_month" in fixed else None

        if product_id not in products:
            products[product_id] = Product(
                product_id=product_id,
                name=product_name or product_id,
                avg_qty_per_order=avg_qty,
                avg_orders_per_month=avg_orders,
                unit_price=unit_price,
            )

        # Extract đơn hàng từ các cột tháng
        for col_idx, month, col_year in month_cols:
            cell_val = row[col_idx] if col_idx < len(row) else None
            iso_dates = parse_month_cell(cell_val, month, col_year)
            for iso in iso_dates:
                qty = avg_qty if avg_qty else 1.0
                revenue = qty * unit_price if (qty and unit_price) else None
                orders.append(OrderEvent(
                    customer_id=customer_id,
                    product_id=product_id,
                    order_date=iso,
                    qty=qty,
                    estimated_revenue=revenue,
                ))

    return customers, products, orders, year


def parse_care_history_sheet(ws, customers: dict[str, Customer]):
    """Sheet 2: merge vào Customer.care_*."""
    # Tìm header row
    for header_row_idx in range(1, min(10, ws.max_row + 1)):
        row = [ws.cell(row=header_row_idx, column=c).value for c in range(1, ws.max_column + 1)]
        normalized = [normalize_header(c) for c in row]
        if "ma kh" in normalized or "ten kh" in normalized:
            break
    else:
        return

    colmap: dict[str, int] = {}
    for idx, h in enumerate(row):
        n = normalize_header(h)
        if n == "ma kh": colmap["id"] = idx
        elif n == "ten kh": colmap["name"] = idx
        elif "muc tieu cham soc" in n: colmap["goal"] = idx
        elif "noi dung tin nhan" in n: colmap["msg"] = idx
        elif "lich su cham soc" in n: colmap["history"] = idx

    for r in range(header_row_idx + 1, ws.max_row + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if all(v is None for v in vals):
            continue
        ma_kh = safe_str(vals[colmap["id"]]) if "id" in colmap else None
        if not ma_kh or ma_kh not in customers:
            continue
        cust = customers[ma_kh]
        if "goal" in colmap:
            cust.care_goal = safe_str(vals[colmap["goal"]]) or cust.care_goal
        if "msg" in colmap:
            cust.care_message_template = safe_str(vals[colmap["msg"]]) or cust.care_message_template
        if "history" in colmap:
            cust.care_history = safe_str(vals[colmap["history"]]) or cust.care_history


def parse_customer_list_sheet(ws, customers: dict[str, Customer]):
    """Sheet 3: merge opportunity/action/status theo tên hoặc SĐT."""
    for header_row_idx in range(1, min(10, ws.max_row + 1)):
        row = [ws.cell(row=header_row_idx, column=c).value for c in range(1, ws.max_column + 1)]
        normalized = [normalize_header(c) for c in row]
        if "ten khach hang" in normalized or "ten kh" in normalized:
            break
    else:
        return

    colmap: dict[str, int] = {}
    for idx, h in enumerate(row):
        n = normalize_header(h)
        if n in ("ten khach hang", "ten kh"): colmap["name"] = idx
        elif "so dien thoai" in n or n == "sdt": colmap["phone"] = idx
        elif "dia chi" in n: colmap["address"] = idx
        elif "hien trang cham soc" in n: colmap["status"] = idx
        elif "co hoi" in n: colmap["opp"] = idx
        elif "can lam" in n: colmap["action"] = idx

    # Lookup: name lowercase -> customer; phone -> customer
    by_name = {c.name.lower(): c for c in customers.values() if c.name}
    by_phone = {c.phone: c for c in customers.values() if c.phone}

    for r in range(header_row_idx + 1, ws.max_row + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if all(v is None for v in vals):
            continue
        name = safe_str(vals[colmap["name"]]) if "name" in colmap else None
        phone = safe_str(vals[colmap["phone"]]) if "phone" in colmap else None

        cust = None
        if name and name.lower() in by_name:
            cust = by_name[name.lower()]
        elif phone and phone in by_phone:
            cust = by_phone[phone]

        if cust is None:
            continue

        if "status" in colmap:
            cust.status = safe_str(vals[colmap["status"]]) or cust.status
        if "opp" in colmap:
            cust.opportunity = safe_str(vals[colmap["opp"]]) or cust.opportunity
        if "action" in colmap:
            cust.action_needed = safe_str(vals[colmap["action"]]) or cust.action_needed


def parse_weca_file(path: str) -> ParsedData:
    """Parse file WECA đa sheet -> ParsedData."""
    wb = load_workbook(path, data_only=True)
    warnings: list[str] = []

    main_ws = None
    care_ws = None
    list_ws = None

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        kind = classify_sheet(ws)
        if kind == "main" and main_ws is None:
            main_ws = ws
        elif kind == "care_history" and care_ws is None:
            care_ws = ws
        elif kind == "customer_list" and list_ws is None:
            list_ws = ws

    if main_ws is None:
        raise ValueError("Không tìm thấy sheet chính (bảng ma trận tháng) trong file WECA.")

    customers, products, orders, year = parse_main_sheet(main_ws, warnings)
    if care_ws is not None:
        parse_care_history_sheet(care_ws, customers)
    else:
        warnings.append("Không có sheet 'Lịch sử chăm sóc' — bỏ qua care_*.")
    if list_ws is not None:
        parse_customer_list_sheet(list_ws, customers)
    else:
        warnings.append("Không có sheet danh sách KH — bỏ qua opportunity/action.")

    return ParsedData(
        source_file=os.path.basename(path),
        parsed_at=now_iso(),
        format="weca",
        year=year,
        customers=list(customers.values()),
        products=list(products.values()),
        orders=orders,
        warnings=warnings,
    )


def main():
    ap = argparse.ArgumentParser(description="Parser WECA: Excel matrix -> JSON.")
    ap.add_argument("input", help="Đường dẫn file .xlsx WECA")
    ap.add_argument("--output", "-o", default="parsed.json", help="File JSON đầu ra")
    args = ap.parse_args()

    if not Path(args.input).exists():
        print(f"Lỗi: không tìm thấy file '{args.input}'.", file=sys.stderr)
        sys.exit(1)

    try:
        parsed = parse_weca_file(args.input)
    except Exception as e:
        print(f"Lỗi khi parse file WECA: {e}", file=sys.stderr)
        sys.exit(2)

    parsed.save_json(args.output)
    print(f"OK: {len(parsed.customers)} KH, {len(parsed.products)} SP, {len(parsed.orders)} đơn -> {args.output}")
    if parsed.warnings:
        print("Cảnh báo:")
        for w in parsed.warnings:
            print(f"  - {w}")


if __name__ == "__main__":
    main()
