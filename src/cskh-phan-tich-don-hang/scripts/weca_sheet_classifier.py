"""
Phân loại 3 sheet của file WECA.
Author: Techla — v1.0.0 — License: xem LICENSE.md

Phân loại theo ưu tiên:
  1. care_history — chứa hint 'muc tieu cham soc' / 'lich su cham soc'
  2. customer_list — 'hien trang cham soc' / 'co hoi' / 'can lam'
  3. main — có 'ma kh' + 'ma sp' + có cột tháng MM/YY

Quan trọng: check care_history TRƯỚC main vì sheet care cũng có 'ma kh'.
"""
from __future__ import annotations
import os
import sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl.worksheet.worksheet import Worksheet
from parser_utils import normalize_header
from weca_date_cell_parser import parse_column_header_to_month


def classify_sheet(ws: Worksheet) -> str:
    """Trả 'main' | 'care_history' | 'customer_list' | 'unknown'."""
    # Đọc tối đa 10 hàng đầu, gom mọi cell non-empty -> normalized
    sample_cells = []
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
        for c in row:
            if c is None:
                continue
            sample_cells.append(normalize_header(c))
    if not sample_cells:
        return "unknown"

    joined = " | ".join(sample_cells)

    # Rule 1: care_history (PRIORITY - check first)
    care_hints = ["muc tieu cham soc", "lich su cham soc", "noi dung tin nhan"]
    if any(h in joined for h in care_hints):
        return "care_history"

    # Rule 2: customer_list
    cl_hints = ["hien trang cham soc", "co hoi", "can lam"]
    cl_match = sum(1 for h in cl_hints if h in joined)
    if cl_match >= 2:
        return "customer_list"

    # Rule 3: main — cần có ma kh + ma sp + ít nhất 1 cột tháng
    has_ma_kh = "ma kh" in joined
    has_ma_sp = "ma sp" in joined
    has_month_col = False
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
        for c in row:
            if parse_column_header_to_month(c):
                has_month_col = True
                break
        if has_month_col:
            break

    if has_ma_kh and has_ma_sp and has_month_col:
        return "main"

    # Secondary customer_list: có "hien trang cham soc" hoặc combo ten-sdt-dia chi
    if "hien trang cham soc" in joined:
        return "customer_list"
    if "ten khach hang" in joined and "dia chi" in joined:
        return "customer_list"

    return "unknown"


def find_header_row_main(ws: Worksheet, max_scan: int = 10) -> int | None:
    """Tìm số dòng header của sheet main. Thường là dòng 5."""
    for row_idx in range(1, max_scan + 1):
        row = [ws.cell(row=row_idx, column=c).value for c in range(1, min(ws.max_column + 1, 50))]
        normalized = [normalize_header(c) for c in row if c is not None]
        joined = " | ".join(normalized)
        if "ma kh" in joined and "ma sp" in joined:
            return row_idx
    return None
